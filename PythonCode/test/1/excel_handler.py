
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from python_handler import ExcelConfig


class ExcelHandler:

    def __init__(self,fileName):
        self.fileName = fileName

    def openSheet(self,sheetName) -> Worksheet:
        wb = openpyxl.load_workbook(self.fileName)
        sheet = wb[sheetName]
        return sheet

    def getFirstRow(self,sheetName) -> list:
        sheetTitle = []
        sheet = self.openSheet(sheetName)
        row = sheet[1]
        for cell in row:
            sheetTitle.append(cell.value)
        return sheetTitle

    def readSheetAllData(self,sheetName):
        sheet = self.openSheet(sheetName)
        title = self.getFirstRow(sheetName)
        rows = list(sheet.rows)[1:]
        data = []
        for row in rows:
            rowData = []
            for cell in row:
                value = cell.value
                # if isinstance(value,str):
                #     if value.endswith("}"):
                #         # print("A")
                #         value = json.loads(value)
                rowData.append(value)
            data.append(dict(zip(title,rowData)))
            # data.append(rowData)
        return data

    def writeData(self,sheetName,row,col,data):
        wb = openpyxl.load_workbook(self.fileName)
        sheet = wb[sheetName]
        sheet.cell(row,col).value = data
        wb.save(self.fileName)
        wb.close()






if __name__ == '__main__':

    # testData = ExcelHandler(ExcelConfig.excelAbsPath).readSheetAllData(ExcelConfig.outputSheetName)
    # print(testData[0]['段号'])
    # print(len(testData))

    testData = ExcelHandler(ExcelConfig.excelAbsPath).readSheetAllData(ExcelConfig.baseSheetName)
    print(testData[0])
    print(len(testData))
