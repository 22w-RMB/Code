
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler():

    def __init__(self,file):
        self.file = file

    '''
        在函数或方法的后面加 -> 类型，表示此函数返回值是一个 这样的类型
        如下：返回的是一个 Worksheet 类型
    '''
    def open_sheet(self,name) -> Worksheet:
        # 打开工作簿
        wb = openpyxl.load_workbook(self.file)
        sheet = wb[name]
        return sheet

    def getSheetFirstRow(self,name) -> list:
        sheet = self.open_sheet(name)
        row = sheet[1]
        header = []
        for cell in row:
            header.append(cell.value)
        return header

    def readAllData(self,name):
        sheet = self.open_sheet(name)
        rows = list(sheet.rows)[1:]

        data = []

        for row in rows:
            rowData = []
            for cell in row:
                value = cell.value
                # if type(value) is str and value.endswith("}"):
                    # value = eval(,cell.value)
                # if type(value) is str:
                #     if value.endswith("}"):
                #         value = eval(cell.value)
                # print(type(value))
                rowData.append(value)
            data.append(dict(zip(self.getSheetFirstRow(name), rowData)))

        return data


    @staticmethod
    def write(file, sheetName, row, column, data):
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheetName]
        sheet.cell(row,column).value = data

        wb.save(file)

        wb.close()

if __name__ == '__main__':
    excel = ExcelHandler(r"D:\code\PythonCode\Project1\excelRWTest\cases.xlsx")
    sheet = excel.open_sheet("登录")
    print(sheet.cell(1, 1).value)
    print(excel.getSheetFirstRow("登录"))
    print(excel.readAllData("登录"))
    import json

    print(json.loads(excel.readAllData("登录")[0]['json']))
