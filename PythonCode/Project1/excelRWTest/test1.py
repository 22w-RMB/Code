


'''

    python 操作 Excel， 工具
    1、openpyxl：支持 xlsx 新型格式的读写，读取速度还可以
    2、tablib：支持多种格式读写，xls、xlsx、csv、json、yaml、html、pd
    3、xlrd：经典的 Excel 读取库
    4、pandas：功能强，太臃肿了

    openpyxl：
    一、获取表单
        1、wb.active
        2、通过索引，wb.worksheets[索引]
        3、通过sheet 名字，wb['sheet_name']


'''

import openpyxl

# 读取 Excel 文件夹
# 读取文件之前一定要关闭该文件
# windows 下面的路径有反斜杠
wb = openpyxl.load_workbook(r"D:\code\PythonCode\Project1\excelRWTest\cases.xlsx")
print(wb)

# 不直接去获取 _sheets 属性，_的成为私有属性
print(wb._sheets)

# active 是表示被激活，被选择的 sheet
active_sheet = wb.active

# sheetnames 和 _sheets 有什么区别?
# sheeetnames 列表中存储的是字符串， _sheets 里面存储的是对象

# 获取所有表单的正确用法
ws = wb.worksheets
print(ws)

# 获取某一个表单
# 1、通过索引获取
print(wb.worksheets[0])
# 2、通过表单名称获取
print(wb.get_sheet_by_name('Sheet1'))  # 过时
sheet = wb['Sheet1']  # 通过该方式获取 sheet 对象在 Pycharm 中不支持只能提示
print(wb['Sheet1'])


# 读取单个单元格，行和列
# 行和列是从 1 开始的
# 获取单元格对象
cell = sheet.cell(2, 3)
# 获取 cell 的值
print(cell.value)
print(cell.column)

# 获取某一行
print(sheet[1])
# 读取值
for column in sheet[1]:
    print(column)
# 获取某一列
print(sheet['A'])
# 获取多行，切片方式，左闭右闭
print(sheet[1:3])
# 获取多列，切片方式，左闭右闭
# print(sheet["A":"B"])

# # 获取所有数据
total_data = list(sheet.rows)[1:]
print(total_data[1][3].value)
# # 结果：
# # [(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>),
# #  (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>),
# #  (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>)]
#
# for row in total_data:
#     for cell in row:
#         print(cell.value)
#         # 写入
#         cell.value = 1
#
# # 保存 , 提供参数，参数是文件路径
# wb.save(r"D:\code\PythonCode\Project1\excelRWTest\cases.xlsx")